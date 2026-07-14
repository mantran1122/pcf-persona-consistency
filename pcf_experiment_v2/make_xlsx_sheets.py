"""Convert the blinded rating CSVs to formatted .xlsx with dropdown validation."""

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

LABELS = ("none,role_breaking,style_drift,memory_failure,context_confusion,"
          "self_contradiction,domain_drift,meta_disclosure,constraint_violation")

for rater in ("rater1", "rater2"):
    df = pd.read_csv(f"outputs/human_annotations/human_evaluation_sheet_{rater}.csv")
    out = f"outputs/human_annotations/human_evaluation_sheet_{rater}.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="cham_diem")
        ws = w.sheets["cham_diem"]
        n = len(df) + 1
        widths = {"A": 10, "B": 10, "C": 10, "D": 55, "E": 75,
                  "F": 6, "G": 6, "H": 6, "I": 6, "J": 6, "K": 20, "L": 30}
        for col, wd in widths.items():
            ws.column_dimensions[col].width = wd
        wrap = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2, max_row=n):
            for c in row:
                c.alignment = wrap
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDEBF7")
        ws.freeze_panes = "F2"

        dv_score = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
        dv_cf = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
        dv_lab = DataValidation(type="list", formula1=f'"{LABELS}"', allow_blank=True)
        for dv in (dv_score, dv_cf, dv_lab):
            ws.add_data_validation(dv)
        for col in "FGHI":
            dv_score.add(f"{col}2:{col}{n}")
        dv_cf.add(f"J2:J{n}")
        dv_lab.add(f"K2:K{n}")
    print("written", out)
